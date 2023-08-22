from django.contrib.auth.decorators import login_required, user_passes_test
from django.shortcuts import render, redirect, redirect, get_object_or_404
from lmsapp.models import Package, User,DropPickZone, Warehouse
from django.contrib import messages
from django.shortcuts import render, redirect
from lmsapp.forms import ChangePasswordForm, PackageForm, ExcelUploadForm
from django.contrib.auth import update_session_auth_hash
from django.core.mail import send_mail
from django.conf import settings
from django.views.decorators.clickjacking import xframe_options_exempt
import random
import string
from datetime import datetime, timedelta
from django.core.mail import send_mail, BadHeaderError
from django.db.models import Q
import pandas as pd
from django.http import HttpResponse
from lmsapp.utils import send_sms

from openpyxl import load_workbook
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.db import IntegrityError


"""  
The view displays a warehouse dashboard template where warehouse users can see packages grouped by drop_pick_zone, 
select packages and assign them to available couriers
"""
# @login_required
# def warehouse_dashboard(request):
#     greeting_message = get_time_of_day()
#     # Retrieve the current warehouse user
#     warehouse_user = request.user

#     # Retrieve the associated warehouse for the warehouse user
#     warehouse = warehouse_user.warehouse

#     # Retrieve the drop_pick_zones belonging to the warehouse
#     drop_pick_zones = DropPickZone.objects.filter(warehouse=warehouse)

#     # Create a dictionary to store packages by location
#     packages_by_location = {}

#     # Iterate over the drop_pick_zones belonging to the warehouse
#     for drop_pick_zone in drop_pick_zones:
#         # Retrieve the packages dropped off at each drop_pick_zone
#         packages = Package.objects.filter(dropOffLocation=drop_pick_zone, status='dropped_off')

#         # Add the drop_pick_zone and associated packages to the dictionary
#         packages_by_location[drop_pick_zone.name] = packages

#     if request.method == 'POST':
#         selected_packages = request.POST.getlist('selected_packages')
#         courier_id = request.POST.get('courier')

#         if selected_packages and courier_id:
#             courier = get_object_or_404(User, id=courier_id, role='courier', status='available')

#             # Update the packages with the assigned courier and change their status
#             packages = Package.objects.filter(id__in=selected_packages)
#             packages.update(courier=courier, status='dispatched')

#             courier_status = Package.objects.filter(courier=courier, status='dispatched').exists()

#             if courier_status:
#                 courier.status = 'on-trip'
            
#             courier.save()

#             messages.success(request, 'Packages successfully assigned to courier.')

#             return redirect('warehouse_dashboard')

#     available_couriers = User.objects.filter(role='courier', status='available')
#     context = {
#         'packages_by_location': packages_by_location,
#         'greeting_message': greeting_message,
#         'available_couriers': available_couriers
#     }

#     return render(request, 'warehouse/warehouse_dashboard.html', context)

def is_warehouse_user(user):
    return user.role == 'warehouse'

@login_required
@user_passes_test(is_warehouse_user)
def warehouse_dashboard(request):
    packages = Package.objects.filter(
        status='dropped_off',
        dropOffLocation__warehouse=request.user.warehouse
    ).select_related('dropOffLocation').order_by('dropOffLocation__tag')

    if request.method == 'POST':
        selected_packages = request.POST.getlist('selected_packages')
        courier_id = request.POST.get('selectCourier')  # Change the name attribute of the courier select field to "selectCourier"
        if selected_packages and courier_id:
            courier = get_object_or_404(User, id=courier_id, role='courier', status='available')
            packages = Package.objects.filter(id__in=selected_packages)
            
            for package in packages:
                if package.deliveryType == 'premium':
                    package.status = 'en_route'
                elif package.deliveryType == 'express':
                    package.status = 'in_transit'
                else:
                    package.status = 'dispatched'
                package.courier = courier
                package.save()

            courier_status = Package.objects.filter(courier=courier, status='dispatched').exists()

            if courier_status:
                courier.status = 'on-trip'
                courier.save()

            messages.success(request, 'Packages successfully assigned to courier.')
            return redirect('warehouse_dashboard')

    # Filter available couriers based on the warehouse
    available_couriers = User.objects.filter(
        role='courier', status='available', warehouse=request.user.warehouse
    )

    context = {
        'packages': packages,
        'available_couriers': available_couriers
    }

    return render(request, 'warehouse/warehouse_dashboard.html', context)


@login_required
@user_passes_test(is_warehouse_user)
def premium_dashboard(request):
    packages = Package.objects.filter(
        Q(deliveryType='premium'),
        status='upcoming',
        warehouse=request.user.warehouse
    ).select_related('dropOffLocation').order_by('dropOffLocation__tag')

    if request.method == 'POST':
        selected_packages = request.POST.getlist('selected_packages')
        courier_id = request.POST.get('selectCourier')
        if selected_packages and courier_id:
            courier = get_object_or_404(User, id=courier_id, role='courier', status='available')
            packages = Package.objects.filter(id__in=selected_packages)

            for package in packages:
                if package.deliveryType == 'premium':
                    package.status = 'en_route'
                elif package.deliveryType == 'express':
                    package.status = 'in_transit'
                else:
                    package.status = 'dispatched'
                package.courier = courier
                package.save()

            courier_status = Package.objects.filter(courier=courier, status='dispatched').exists()

            if courier_status:
                courier.status = 'on-trip'
                courier.save()

            messages.success(request, 'Packages successfully assigned to courier.')
            return redirect('warehouse_dashboard')

    # Filter couriers based on the warehouse
    available_couriers = User.objects.filter(
        role='courier', status='available', warehouse=request.user.warehouse
    )

    context = {
        'packages': packages,
        'available_couriers': available_couriers
    }

    return render(request, 'warehouse/premium_dashboard.html', context)


@login_required
@user_passes_test(is_warehouse_user)
def express_dashboard(request):
    packages = Package.objects.filter(
        Q(deliveryType='express'),
        status='upcoming',
        warehouse=request.user.warehouse
    ).select_related('dropOffLocation').order_by('dropOffLocation__tag')

    if request.method == 'POST':
        selected_packages = request.POST.getlist('selected_packages')
        courier_id = request.POST.get('selectCourier')  # Change the name attribute of the courier select field to "selectCourier"
        if selected_packages and courier_id:
            courier = get_object_or_404(User, id=courier_id, role='courier', status='available')
            packages = Package.objects.filter(id__in=selected_packages)
            
            for package in packages:
                if package.deliveryType == 'premium':
                    package.status = 'en_route'
                elif package.deliveryType == 'express':
                    package.status = 'in_transit'
                else:
                    package.status = 'dispatched'
                package.courier = courier
                package.save()

            courier_status = Package.objects.filter(courier=courier, status='dispatched').exists()

            if courier_status:
                courier.status = 'on-trip'
                courier.save()

            messages.success(request, 'Packages successfully assigned to courier.')
            return redirect('warehouse_dashboard')

    available_couriers = User.objects.filter(
        role='courier', status='available', warehouse=request.user.warehouse
    )

    context = {
        'packages': packages,
        'available_couriers': available_couriers
    }

    return render(request, 'warehouse/express_dashboard.html', context)



# def warehouse_dashboard(request):
#     # Assuming you have the warehouse information from the user or some other source
#     # Replace this with the correct warehouse ID
#     packages = Package.objects.all()

#     context = {
#         'packages': packages,
#     }

#     return render(request, 'warehouse/warehouse_dashboard.html', context)


"""
The view handles the change password functionality and renders a template with a form 
for users to enter their new password.
"""
@login_required
def change_password(request):
    if request.method == 'POST':
        form = ChangePasswordForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)  # Important to update the session
            messages.success(request, 'Your password has been successfully changed.')
            return redirect('warehouse_dashboard')
        else:
            messages.error(request, 'Please correct the error below.')
    else:
        form = ChangePasswordForm(request.user)
    return render(request, 'warehouse/change_password.html', {'form': form})

""" 
Handles the confirmation of package arrival at the warehouse. It updates the package status, 
updates the courier's status if applicable, sends an email notification to the sender
"""
@login_required
@user_passes_test(is_warehouse_user)
def confirm_arrival(request, package_id):
    if request.method == 'POST':
        warehouse_user = request.user
        warehouse = warehouse_user.warehouse

        package = Package.objects.get(pk=package_id)
        package.status = 'in_house'
        courier = package.courier
        if courier:
            courier.status = 'available'
            courier.save()
        
        # Assign the warehouse to the package
        package.warehouse = warehouse
        
        # Save the package
        package.save()
        
        # Send email to sender
        subject = 'Package Dropped Off at Warehouse'
        message = f'Dear Customer, your package with delivery number {package.package_number} has been dropped off at the warehouse.'

        sender_user = User.objects.get(username=package.user.username)
        sender_email = sender_user.email

        send_mail(subject, message, settings.EMAIL_HOST_USER, [sender_email])   

        messages.success(request, "Package arrival notified successfully.")
    else:
        messages.error(request, "Invalid request.")

    return redirect('in_house')  # Replace with the appropriate URL for the warehouse dashboard

""" 
The view displays a list of packages with the 'in_house'status and allows the user to assign selected packages 
to available couriers and drop_pick_zones, and updates the package status to 'in_transit'.
""" 
@login_required
@user_passes_test(is_warehouse_user)
def in_house(request):
    ready_packages = Package.objects.filter(status='in_house')

    if request.method == 'POST':
        selected_packages = request.POST.getlist('selected_packages')
        courier_id = request.POST.get('courier')

        if selected_packages and courier_id:
            courier = get_object_or_404(User, id=courier_id, role='courier')

            # Update the packages with the assigned courier and change their status
            packages = Package.objects.filter(id__in=selected_packages)
            package_types = set(package.deliveryType for package in packages)
            status_updated = False

            for package in packages:
                if package.deliveryType == 'premium':
                    package.status = 'en_route'
                    status_updated = True
                elif package.deliveryType == 'express':
                    package.status = 'in_transit'
                    status_updated = True
                else:
                    package.status = 'in_transit'  # Change status to 'in_transit' for standard packages
                    status_updated = True
                package.courier = courier
                package.save()

            courier.status = 'on-trip'
            courier.save()

            if status_updated:
                if 'premium' in package_types and 'express' in package_types:
                    messages.success(request, 'Premium and Express packages successfully assigned to courier.')
                elif 'premium' in package_types:
                    messages.success(request, 'Premium packages successfully assigned to courier.')
                elif 'express' in package_types:
                    messages.success(request, 'Express packages successfully assigned to courier.')
                else:
                    messages.success(request, 'Standard packages successfully assigned to courier.')
            else:
                messages.error(request, 'No packages were assigned.')

            return redirect('in_house')
        else:
            messages.error(request, 'Please select packages and a courier.')

    context = {
        'ready_packages': ready_packages,
        'available_couriers': User.objects.filter(role='courier', status='available'),
    }
    return render(request, 'warehouse/ready_packages.html', context)



def to_pickup(request, package_id):
    if request.method == 'POST':
        package = Package.objects.get(pk=package_id)
        package.status = 'in_transit'
        package.save()

    return redirect('ready_packages')  # Replace with the appropriate URL for the warehouse dashboard

""" 
The view displays a list of packages ready for pickup and allows the user to assign selected packages 
to couriers and drop_pick_zones. It handles form submissions, updates the package assignments and statuses, 
and provides available couriers and drop_pick_zones in the context.
""" 
@login_required
@user_passes_test(is_warehouse_user)
def ready_for_pickup(request):
    if request.method == 'POST':
        selected_packages = request.POST.getlist('selected_packages')
        courier_id = request.POST.get('courier')

        if selected_packages and courier_id:
            courier = get_object_or_404(User, id=courier_id, role='courier')

            # Update the packages with the assigned courier and change their status
            packages = Package.objects.filter(id__in=selected_packages, deliveryType=['premium', 'standard'])
            packages.update(courier=courier, status='in_transit')
            
            courier.status = 'on-trip'
            courier.save()

            messages.success(request, 'Premium packages successfully assigned to courier.')

            return redirect('in_house')

    ready_packages = Package.objects.filter(status='in_transit', deliveryType='premium')
    available_couriers = User.objects.filter(role='courier', status='available')

    context = {
        'ready_packages': ready_packages,
        'available_couriers': available_couriers
    }
    return render(request, 'warehouse/ready_for_pickup.html', context)


@login_required
@user_passes_test(is_warehouse_user)
def new_arrivals(request):
    arrived_packages = Package.objects.filter(status='warehouse_arrival')

    if request.method == 'POST':

        return redirect('ready_packages')
            
    context = {
        'arrived_packages': arrived_packages,
    }
    return render(request, 'warehouse/new_arrivals.html', context)

def generate_package_number():
    prefix = 'pn'
    digits = ''.join(random.choices(string.digits, k=5))
    return f'{prefix}{digits}'

# def is_warehouse_user(user):
#     return user.role == 'warehouse'

@login_required
@xframe_options_exempt 
@user_passes_test(is_warehouse_user)
# def add_package(request):

#     senders = User.objects.filter(role='sender')
#     if request.method == 'POST':
#         form = PackageForm(request.POST)

#         if form.is_valid():
#             package = form.save(commit=False)
#             package.user = request.user
#             package.package_number = generate_package_number()

#             # Automatically set the warehouse to the one the user belongs to
#             user_warehouse = request.user.warehouse
#             if user_warehouse:
#                 package.warehouse = user_warehouse
#             else:
#                 # Handle the case where the user does not have a drop_pick_zone
#                 pass

#             recipient_latitude = request.POST.get('recipient_latitude')
#             recipient_longitude = request.POST.get('recipient_longitude')
#             package.recipient_latitude = recipient_latitude
#             package.recipient_longitude = recipient_longitude

#             package.status = 'in_house'
#             package.save()

#             return redirect('warehouse_dashboard')  # Redirect to a success page or wherever you want

#     else:
#         form = PackageForm()
#         user_warehouse = request.user.warehouse  # Get the user's drop_pick_zone

#     # Get the drop_pick_zones data to populate the recipientPickUpLocation dropdown
#     warehouse = Warehouse.objects.all()  # Adjust this based on your model
#     context = {'form': form, 'warehouse': warehouse, 'user_warehouse': user_warehouse, 'senders': senders }
#     return render(request, 'warehouse/add_package.html', context)
@login_required
@xframe_options_exempt 
@user_passes_test(is_warehouse_user)

def add_package(request):
    msg = None
    user_warehouse = None  # Initialize user_warehouse
    
    senders = User.objects.filter(role='sender')
    if request.method == 'POST':
        form = PackageForm(request.POST)

        if form.is_valid():
            package = form.save(commit=False)
            package.created_by = request.user
            package.package_number = generate_package_number()

            # Automatically set the warehouse to the one the user belongs to
            user_warehouse = request.user.warehouse
            if user_warehouse:
                package.warehouse = user_warehouse
            else:
                # Handle the case where the user does not have a warehouse
                pass

            recipient_latitude = request.POST.get('recipient_latitude')
            recipient_longitude = request.POST.get('recipient_longitude')
            package.recipient_latitude = recipient_latitude
            package.recipient_longitude = recipient_longitude

            package.status = 'in_house'

            selected_user_id = request.POST.get('user')
            if selected_user_id:
                selected_user = User.objects.get(id=selected_user_id)

                package.sendersEmail = selected_user.email
                package.sendersName = selected_user.username
            package.save()

            
            # if 'userCheckbox' in request.POST:
            #     selected_user_id = request.POST.get('user')
            #     if selected_user_id:
            #         selected_user = User.objects.get(id=selected_user_id)
            #         print(selected_user)
            #         package.sendersEmail = selected_user.email
            #         package.sendersName = selected_user.username

            # # Save the package after updating the fields
            # # print("Form Data:", request.POST)  # Print the form data
            # package.save()


            # Send an email to the sender
            subject = 'Package Registered'
            message = (
                f"Dear Customer, your package has been successfully registered.\n\n"
                f"Package Number: {package.package_number}\n"
                f"Recipient: {package.recipientName}\n"
                f"Recipient Address: {package.recipientAddress}\n"
                f"If you have any questions, please contact our customer support team.\n"
            )


            if package.dropOffLocation:
                message += f"Drop-off Location: {package.dropOffLocation.name}\n"

            message += f"Status: {package.status}"
            
            from_email = settings.DEFAULT_FROM_EMAIL
            recipient_list = [package.recipientEmail, package.sendersEmail]
                           
            sender_contact = str(package.sendersContact).strip()

            if len(sender_contact) == 10 and sender_contact.startswith('0'):
                sender_contact = '+256' + sender_contact[1:]

            send_sms([sender_contact], message, settings.AFRICASTALKING_SENDER)
           
            try:
                send_mail(subject, message, from_email, recipient_list, fail_silently=False)
            except BadHeaderError as e:
                # Handle the BadHeaderError
                print("Error: BadHeaderError:", str(e))
            except Exception as e:
                # Handle other exceptions
                print("Error:", str(e))

            msg = 'Package registered'
            return redirect('in_house')  # Redirect to a success page or wherever you want
        else:
            msg = 'Form is not valid'
    else:
        form = PackageForm()
        user_warehouse = request.user.warehouse

    # Get the warehouse data to populate the warehouse dropdown
    warehouse = Warehouse.objects.all()  # Adjust this based on your model
    context = {'form': form, 'warehouse': warehouse, 'user_warehouse': user_warehouse, 'senders': senders}
    return render(request, 'warehouse/add_package.html', context)

@login_required
@user_passes_test(is_warehouse_user)
def warehouse_reports(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    packages_delivered = Package.objects.filter(status='completed')
    packages_ready = Package.objects.filter(status='ready_for_pickup')
    
    if start_date and end_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
        
        # Ensure that start_date is always before end_date
        if start_date > end_date:
            start_date, end_date = end_date, start_date
        
        # Adjust end_date to be inclusive of the entire day
        end_date += timedelta(days=1)
        
        packages_delivered = packages_delivered.filter(completed_at__range=(start_date, end_date))
    elif start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        packages_delivered = packages_delivered.filter(completed_at__date=start_date)
        
    
    context = {
        'packages_delivered': packages_delivered,
        'packages_ready': packages_ready,
    }
    
    return render(request, 'warehouse/warehouse_reports.html', context )

@login_required
@user_passes_test(is_warehouse_user)
def package_reports_export(request):
    start_datetime = request.GET.get('start_datetime')
    end_datetime = request.GET.get('end_datetime')

    packages_delivered = Package.objects.filter(status='completed')
    # packages_ready = Package.objects.filter(status='ready_for_pickup')
    
    if start_datetime and end_datetime:
        start_datetime = datetime.strptime(start_datetime, '%Y-%m-%dT%H:%M')
        end_datetime = datetime.strptime(end_datetime, '%Y-%m-%dT%H:%M')
        
        # Ensure that start_datetime is always before end_datetime
        if start_datetime > end_datetime:
            start_datetime, end_datetime = end_datetime, start_datetime
        
        packages_delivered = packages_delivered.filter(completed_at__range=(start_datetime, end_datetime))
    elif start_datetime:
        start_datetime = datetime.strptime(start_datetime, '%Y-%m-%dT%H:%M')
        packages_delivered = packages_delivered.filter(completed_at__gte=start_datetime)
        
    context = {
        'packages_delivered': packages_delivered,
        # 'packages_ready': packages_ready,
    }

    # Get the current date and time
    current_datetime = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')

    # Create a Pandas DataFrame from the data
    df = pd.DataFrame(context)

    # Create a response with the Excel file
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment; filename="Package_summary_data_{current_datetime}.xlsx"'

    # Save the DataFrame to the Excel response
    df.to_excel(response, index=False, engine='openpyxl')
    
    return response

# @login_required
# @user_passes_test(is_warehouse_user)
# def upload_excel(request):
#     username_list = ['Ivan', 'muhumuza', 'Shem', 'Izzy']

#     senders = User.objects.filter(username__in=username_list, role='sender')

#     if request.method == "POST":
#         form = ExcelUploadForm(request.POST, request.FILES)
#         if form.is_valid():
#             excel_file = form.cleaned_data['excel_file']
#             wb = load_workbook(excel_file)
#             sheet = wb.active

#             header_mapping = {
#                 "order_id":1, 
#                 "order_date":2, 
#                 "recipient_name":3, 
#                 "delivery_address":4, 
#                 "city":5, 
#                 "phone":6, 
#                 "item" :7
#             }

#             for row in sheet.iter_rows(min_row=2, values_only=True):
                
#                 # Format the order_date value to a string in ISO format
#                 order_date_str = row[header_mapping["order_date"]].strftime("%Y-%m-%d")

#                  # Check if the phone field is blank
#                 recipient_telephone = row[header_mapping["phone"]]
#                 if not recipient_telephone:
#                     recipient_telephone = "00000"  

#                 package = Package(
#                     user=request.POST.get('user'),
#                     packageName=row[header_mapping["item"]],
#                     deliveryType='premium',  
#                     package_number=row[header_mapping["order_id"]],
#                     recipientName=row[header_mapping["recipient_name"]],
#                     recipientEmail='',  
#                     recipientTelephone=recipient_telephone,
#                     recipientAddress=row[header_mapping["delivery_address"]] + ', ' + row[header_mapping["city"]],
#                     packageDescription='',
#                     sendersName='Sam Kulubya',  
#                     sendersEmail='selmo@gmail.com',  
#                     sendersAddress='',  
#                     sendersContact='',  
#                     created_on=parse_date(order_date_str),
#                     created_by=request.user,
#                     modified_by=request.user,
#                     assigned_at=timezone.now(),
#                     status='warehouse_arrival',
#                     warehouse=request.user.warehouse
#                 )

#                 try:
#                     package.save()
#                 except IntegrityError as ie:
#                     if "UNIQUE constraint" in str(ie):
#                         # Handle integrity error when package_number is already taken
#                         messages.error(request, "Error: The package number already exists.")
#                     else:
#                         messages.error(request, "An error occurred while saving the package.")

#             return redirect("new_arrivals")
#     else:
#         form = ExcelUploadForm()

#     return render(request, "warehouse/upload_excel.html", {"form": form, "senders": senders})

def preprocess_phone_number(phone_number):
    # Convert phone_number to string if it's an integer
    phone_number_str = str(phone_number)
    
    # Remove non-digit characters and return the cleaned phone number
    cleaned_phone_number = ''.join(filter(str.isdigit, phone_number_str))

    # Add the leading zero if missing
    if not cleaned_phone_number.startswith('0'):
        cleaned_phone_number = '0' + cleaned_phone_number

    return cleaned_phone_number

from django.db import IntegrityError
from openpyxl import load_workbook
# from django.utils.timezone import parse_date
from django.utils.dateparse import parse_date
from django.contrib import messages
from django.shortcuts import render, redirect
from django.utils import timezone
from lmsapp.models import Package

@login_required
@user_passes_test(is_warehouse_user)
def upload_excel(request):
    username_list = ['Ivan', 'muhumuza', 'Shem', 'Izzy']

    senders = User.objects.filter(username__in=username_list, role='sender')

    if request.method == "POST":
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = form.cleaned_data['excel_file']
            wb = load_workbook(excel_file)
            sheet = wb.active

            header_mapping = {
                "order_id": 1, 
                "order_date": 2, 
                "recipient_name": 3, 
                "delivery_address": 4, 
                "city": 5, 
                "phone": 6, 
                "item": 7
            }

            for row in sheet.iter_rows(min_row=2, values_only=True):

                order_date_value = row[header_mapping["order_date"]]   
                if order_date_value is None:
                    # Handle the case where order_date_value is None (blank cell)
                    continue
                
                order_date_str = order_date_value.strftime("%Y-%m-%d")

                # Check if the phone field is blank
                recipient_telephone = preprocess_phone_number(row[header_mapping["phone"]])
                if not recipient_telephone:
                    recipient_telephone = "Not Provided"  

                recipient_address = row[header_mapping["delivery_address"]]
                city = row[header_mapping["city"]]

                if not recipient_address and not city:
                    recipient_address_city = ""
                elif not recipient_address:
                    recipient_address_city = city
                elif not city:
                    recipient_address_city = recipient_address
                else:
                    recipient_address_city = recipient_address + ', -' + city

                package_number = row[header_mapping["order_id"]]
                order_date = parse_date(order_date_str)

                # Check if a package with the same package_number already exists
                existing_package = Package.objects.filter(package_number=package_number).first()

                if existing_package:
                    # Package with the same package_number already exists, skip
                    continue
                
                
                user_id = request.POST.get('client')
                try:
                    user = User.objects.get(id=user_id)
                except User.DoesNotExist:
                    user = None

                # Check if user has the 'name' attribute, if not, use 'username'
                if user and hasattr(user, 'name') and user.name:
                    senders_name = user.name
                else:
                    senders_name = user.username

                package = Package(
                    user=user,
                    packageName=row[header_mapping["item"]],
                    deliveryType='premium',  
                    package_number=package_number,
                    recipientName=row[header_mapping["recipient_name"]],
                    recipientEmail='',  
                    recipientTelephone=recipient_telephone,
                    recipientAddress=recipient_address_city,
                    packageDescription='',
                    sendersName=senders_name,  
                    sendersEmail=user.email,  
                    sendersAddress=user.address,  
                    sendersContact='',  
                    created_on=order_date,
                    created_by=user,
                    modified_by=request.user,
                    assigned_at=timezone.now(),
                    status='warehouse_arrival',
                    warehouse=request.user.warehouse
                )

                try:
                    package.save()
                except IntegrityError as ie:
                    if "UNIQUE constraint" in str(ie):
                        # Handle integrity error when package_number is already taken
                        messages.error(request, "Error: The package number already exists.")
                    else:
                        messages.error(request, "An error occurred while saving the package.")

            return redirect("new_arrivals")
    else:
        form = ExcelUploadForm()

    return render(request, "warehouse/upload_excel.html", {"form": form, "senders": senders})





# import gspread
# from oauth2client.service_account import ServiceAccountCredentials

# scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive',
#          'https://www.googleapis.com/auth/drive.file', 'https://www.googleapis.com/auth/spreadsheets']


# credentials = ServiceAccountCredentials.from_json_keyfile_name('cred.json', scope)

# # creds = ServiceAccountCredentials.from_json_keyfile_name('mod.json', scope)
# client = gspread.authorize(credentials)

# sheet = client.open_by_url("https://docs.google.com/spreadsheets/d/1xsk1k8F1rJJ4nj350LGSJpRXUIBv_F-t6KPIyGdCJ_I/edit#gid=0").sheet1

# data = sheet.get_all_records()

# # sheet.update_cell(1, 1, "You made it") #Write this message in first row and first column

# print(data)